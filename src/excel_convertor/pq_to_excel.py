import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import re
import os

def create_prequalification_excel(md_file, excel_file):
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pre-Qualification Criteria"
    
    # Extract project title dynamically
    title_match = re.search(r'# (.+)', content)
    project_title = title_match.group(1) if title_match else "Pre-Qualification Criteria"
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = project_title
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 3
    
    # Extract all tables dynamically
    table_pattern = r'\|[^\n]*\|(?:\n\|[^\n]*\|)+'
    tables = re.findall(table_pattern, content)
    
    for i, table in enumerate(tables):
        lines = [line.strip() for line in table.strip().split('\n') if line.strip() and '|' in line]
        if len(lines) < 2:
            continue
            
        # Extract header
        header_line = lines[0]
        headers = [cell.strip() for cell in header_line.split('|')[1:-1] if cell.strip()]
        if not headers:
            continue
            
        # Skip separator lines and extract data
        data_lines = [line for line in lines[1:] if not re.match(r'^\|[-\s|]+\|$', line)]
        
        # Section header based on content
        section_name = f"PRE-QUALIFICATION TABLE {i+1}"
        if any(word in str(headers).lower() for word in ['description', 'mandatory', 'documents']):
            section_name = "PRE-QUALIFICATION CRITERIA"
        elif any(word in str(headers).lower() for word in ['details', 'section', 'checklist']):
            section_name = "EVALUATION CHECKLIST"
        elif any(word in str(headers).lower() for word in ['particulars', 'instructions']):
            section_name = "BID SUBMISSION INSTRUCTIONS"
        elif any(word in str(headers).lower() for word in ['item', 'emd', 'fee', 'validity']):
            section_name = "DEADLINES & REQUIREMENTS"
        
        ws.merge_cells(f'A{current_row}:{chr(64+len(headers))}{current_row}')
        ws[f'A{current_row}'] = section_name
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col_num, value=header)
            ws.cell(row=current_row, column=col_num).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=col_num).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.cell(row=current_row, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Add data
        for line in data_lines:
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            if len(cells) >= len(headers) and any(cell for cell in cells):
                for col_num, cell_value in enumerate(cells, 1):
                    if col_num <= len(headers):
                        ws.cell(row=current_row, column=col_num, value=cell_value)
                        ws.cell(row=current_row, column=col_num).alignment = Alignment(wrap_text=True, vertical='top')
                current_row += 1
        
        current_row += 2
    
    # Extract key text sections
    sections = [
        ("1. GENERAL NOTES", r'## 1\. General Notes\s*(.*?)(?=## 2\.|$)'),
        ("4. REJECTION CRITERIA", r'## 4\. Rejection Criteria Related to PQ\s*(.*?)(?=## 5\.|$)')
    ]
    
    for section_title, pattern in sections:
        section_match = re.search(pattern, content, re.DOTALL)
        if section_match:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = section_title
            ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
            ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
            current_row += 1
            
            section_text = section_match.group(1).strip()
            # Remove table content from text sections
            section_text = re.sub(r'\|[^\n]*\|(?:\n\|[^\n]*\|)+', '', section_text)
            section_lines = [line.strip() for line in section_text.split('\n') if line.strip() and not line.startswith('|')]
            
            for line in section_lines:
                ws.merge_cells(f'A{current_row}:F{current_row}')
                ws[f'A{current_row}'] = line
                ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
                current_row += 1
            
            current_row += 1
    
    # Set dynamic column widths
    for col in range(1, ws.max_column + 1):
        col_letter = chr(64 + col)
        if col == 1:
            ws.column_dimensions[col_letter].width = 8
        elif col == 2:
            ws.column_dimensions[col_letter].width = 50
        else:
            ws.column_dimensions[col_letter].width = 30
    
    wb.save(excel_file)
    print(f"Pre-Qualification Excel file created: {excel_file}")

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python pq_to_excel.py <markdown_file> [excel_file]")
        exit(1)
    
    md_file = sys.argv[1]
    excel_file = sys.argv[2] if len(sys.argv) > 2 else md_file.replace('.md', '.xlsx')
    
    if not os.path.exists(md_file):
        print(f"File not found: {md_file}")
        exit(1)
    
    create_prequalification_excel(md_file, excel_file)
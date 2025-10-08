import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import os

def create_rfp_excel(md_file, excel_file):
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "RFP Key Details"
    
    # Extract project title
    project_match = re.search(r'\*\*Project Title:\*\* (.+)', content)
    project_title = project_match.group(1) if project_match else "RFP Project"
    
    # Header
    ws.merge_cells('A1:B1')
    ws['A1'] = project_title
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Extract table data
    table_match = re.search(r'\| Key Detail \| Information \|\s*\|[-\|]+\|(.*?)(?=## Scope of Work)', content, re.DOTALL)
    
    current_row = 3
    if table_match:
        table_data = table_match.group(1).strip()
        rows = [row.strip() for row in table_data.split('\n') if row.strip() and '|' in row]
        
        # Headers
        ws['A2'] = "Key Detail"
        ws['B2'] = "Information"
        ws['A2'].font = Font(bold=True, color="FFFFFF")
        ws['B2'].font = Font(bold=True, color="FFFFFF")
        ws['A2'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['B2'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for row in rows:
            cells = [cell.strip() for cell in row.split('|')[1:-1]]
            if len(cells) >= 2:
                ws[f'A{current_row}'] = cells[0]
                ws[f'B{current_row}'] = cells[1]
                ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws[f'B{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
                current_row += 1
    
    # Scope of Work section
    current_row += 1
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = "SCOPE OF WORK"
    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{current_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    
    scope_match = re.search(r'## Scope of Work\s*(.*?)(?=## Additional Key Details)', content, re.DOTALL)
    if scope_match:
        scope_text = scope_match.group(1).strip()
        scope_lines = [line.strip('- ').strip() for line in scope_text.split('\n') if line.strip()]
        
        for line in scope_lines:
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws[f'A{current_row}'] = line
            ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1
    
    # Additional Key Details section
    current_row += 1
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = "ADDITIONAL KEY DETAILS"
    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{current_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    
    additional_match = re.search(r'## Additional Key Details\s*(.*?)(?=---)', content, re.DOTALL)
    if additional_match:
        additional_text = additional_match.group(1).strip()
        additional_lines = [line.strip('- ').strip() for line in additional_text.split('\n') if line.strip()]
        
        for line in additional_lines:
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws[f'A{current_row}'] = line
            ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80
    
    wb.save(excel_file)
    print(f"Excel file created: {excel_file}")

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python rfp_summary_to_excel.py <markdown_file> [excel_file]")
        exit(1)
    
    md_file = sys.argv[1]
    excel_file = sys.argv[2] if len(sys.argv) > 2 else md_file.replace('.md', '.xlsx')
    
    if not os.path.exists(md_file):
        print(f"File not found: {md_file}")
        exit(1)
    
    create_rfp_excel(md_file, excel_file)
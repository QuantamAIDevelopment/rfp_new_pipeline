import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import re
import os

def create_tq_excel(md_file, excel_file):
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Technical Qualification"
    
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "Technical Qualification Criteria (Pure Technical Scoring)"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 3
    
    # Technical Evaluation Parameters
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "TECHNICAL EVALUATION PARAMETERS (100 Marks Total)"
    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    
    params = ["Bidder's Competence (80 Marks)", "Quality of Technical Proposal (10 Marks)", "Technical Proposal Understanding (10 Marks)"]
    for param in params:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = param
        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True)
        current_row += 1
    
    current_row += 1
    
    # Technical Qualification Scoring Table
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "TECHNICAL QUALIFICATION SCORING TABLE"
    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    
    # Table headers
    headers = ["Sr. No.", "Technical Qualification Criteria", "Maximum Marks", "Scoring Mechanism", "Supporting Documents"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=header)
        ws.cell(row=current_row, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    # Extract table data
    table_pattern = r'\| Sr\. No\. \| Parameter \| Maximum Marks \| Marks \| Supporting Documents \|(.*?)(?=##|$)'
    table_match = re.search(table_pattern, content, re.DOTALL)
    
    if table_match:
        table_data = table_match.group(1).strip()
        rows = [row.strip() for row in table_data.split('\n') if row.strip() and '|' in row and not re.match(r'^\|[-\s|]+\|$', row)]
        
        for row in rows:
            cells = [cell.strip() for cell in row.split('|')[1:-1]]
            if len(cells) >= 4 and any(cell for cell in cells):
                for col, cell_value in enumerate(cells[:5], 1):
                    ws.cell(row=current_row, column=col, value=cell_value)
                    ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
                current_row += 1
    
    current_row += 1
    
    # Technical Evaluation Process
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "TECHNICAL EVALUATION PROCESS"
    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    
    process_match = re.search(r'## Technical Evaluation Process\s*(.*?)$', content, re.DOTALL)
    if process_match:
        process_text = process_match.group(1).strip()
        process_lines = [line.strip('- ').strip() for line in process_text.split('\n') if line.strip()]
        
        for line in process_lines:
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = line
            ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 30
    
    wb.save(excel_file)
    print(f"Excel file created: {excel_file}")

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python pure_tq_to_excel.py <markdown_file> [excel_file]")
        exit(1)
    
    md_file = sys.argv[1]
    excel_file = sys.argv[2] if len(sys.argv) > 2 else md_file.replace('.md', '.xlsx')
    
    if not os.path.exists(md_file):
        print(f"File not found: {md_file}")
        exit(1)
    
    create_tq_excel(md_file, excel_file)
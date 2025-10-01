import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import re
import os

def create_boq_excel(md_file, excel_file):
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ - Bill of Quantities"
    
    # Extract project title dynamically
    title_match = re.search(r'# (.+)', content)
    project_title = title_match.group(1) if title_match else "Bill of Quantities (BOQ)"
    
    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = project_title
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 3
    
    # Extract all tables dynamically
    table_pattern = r'\|[^\n]*\|(?:\n\|[^\n]*\|)+'
    tables = re.findall(table_pattern, content)
    
    for i, table in enumerate(tables):
        lines = [line.strip() for line in table.strip().split('\n') if line.strip() and '|' in line]
        if len(lines) < 2:
            continue
            
        # Extract header
        header_line = lines[0]
        headers = [cell.strip() for cell in header_line.split('|')[1:-1] if cell.strip()]
        if not headers:
            continue
            
        # Skip separator lines and extract data
        data_lines = [line for line in lines[1:] if not re.match(r'^\|[-\s|]+\|$', line)]
        
        # Section header
        section_name = f"BOQ TABLE {i+1}"
        if i == 0 and any(word in str(headers).lower() for word in ['position', 'manpower', 'resource']):
            section_name = "MANPOWER REQUIREMENTS"
        elif any(word in str(headers).lower() for word in ['cost', 'price', 'amount']):
            section_name = "COST STRUCTURE"
        
        ws.merge_cells(f'A{current_row}:{chr(64+len(headers))}{current_row}')
        ws[f'A{current_row}'] = section_name
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col_num, value=header)
            ws.cell(row=current_row, column=col_num).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=col_num).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.cell(row=current_row, column=col_num).alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Add data
        for line in data_lines:
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            if len(cells) >= len(headers) and any(cell for cell in cells):
                for col_num, cell_value in enumerate(cells, 1):
                    if col_num <= len(headers):
                        ws.cell(row=current_row, column=col_num, value=cell_value)
                        ws.cell(row=current_row, column=col_num).alignment = Alignment(wrap_text=True, vertical='top')
                current_row += 1
        
        current_row += 2
    
    # BOQ Notes Section
    notes_match = re.search(r'## 2\. BOQ Notes / Instructions\s*(.*?)(?=---|$)', content, re.DOTALL)
    if notes_match:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "BOQ NOTES & INSTRUCTIONS"
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
        
        notes_text = notes_match.group(1).strip()
        notes_lines = [line.strip() for line in notes_text.split('\n') if line.strip()]
        
        for note in notes_lines:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = note
            ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1
    
    # Set dynamic column widths
    for col in range(1, ws.max_column + 1):
        col_letter = chr(64 + col)
        if col == 1:
            ws.column_dimensions[col_letter].width = 8
        elif col == 2:
            ws.column_dimensions[col_letter].width = 35
        else:
            ws.column_dimensions[col_letter].width = 20
    
    wb.save(excel_file)
    print(f"BOQ Excel file created: {excel_file}")

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python boq_to_excel.py <markdown_file> [excel_file]")
        exit(1)
    
    md_file = sys.argv[1]
    excel_file = sys.argv[2] if len(sys.argv) > 2 else md_file.replace('.md', '.xlsx')
    
    if not os.path.exists(md_file):
        print(f"File not found: {md_file}")
        exit(1)
    
    create_boq_excel(md_file, excel_file)
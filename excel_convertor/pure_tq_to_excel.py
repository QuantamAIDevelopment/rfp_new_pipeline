import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import re
import os
 
def create_tq_excel(md_file, excel_file):
    with open(md_file, 'r', encoding='utf-8') as f:
        content = f.read()
   
    wb = Workbook()
    ws = wb.active
    ws.title = "Technical Qualification"
   
    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "Technical Qualification Criteria (Pure Technical Scoring)"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
   
    current_row = 3
   
    # Technical Evaluation Parameters section
    eval_params_match = re.search(r'## Technical Evaluation Parameters\s*\n(.*?)(?=##)', content, re.DOTALL)
    if eval_params_match:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = "TECHNICAL EVALUATION PARAMETERS"
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        current_row += 1
       
        # Parse evaluation parameters table
        eval_table = eval_params_match.group(1).strip()
        eval_rows = [row.strip() for row in eval_table.split('\n') if '|' in row and not re.match(r'^[-\s|]+$', row)]
       
        for row in eval_rows:
            cells = [cell.strip() for cell in row.split('|') if cell.strip()]
            if len(cells) >= 3:
                ws.cell(row=current_row, column=1, value=cells[0])
                ws.cell(row=current_row, column=2, value=cells[1])
                ws.cell(row=current_row, column=3, value=cells[2])
                for col in range(1, 4):
                    ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
                current_row += 1
        current_row += 1
   
    # Technical Qualification Scoring Table
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "TECHNICAL QUALIFICATION SCORING TABLE"
    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
   
    # Table headers
    headers = ["#", "Criteria Category", "Evaluation criterion details", "Max Marks", "Supporting Documents required"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=current_row, column=col, value=header)
        ws.cell(row=current_row, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
   
    # Extract main scoring table
    scoring_table_match = re.search(r'## Technical Qualification Scoring Table\s*\n(.*?)(?=##)', content, re.DOTALL)
    if scoring_table_match:
        table_data = scoring_table_match.group(1).strip()
        rows = [row.strip() for row in table_data.split('\n') if '|' in row and not re.match(r'^[-\s|]+$', row)]
       
        for row in rows[1:]:  # Skip header row
            cells = [cell.strip() for cell in row.split('|') if cell.strip()]
            if len(cells) >= 4:
                for col, cell_value in enumerate(cells[:5], 1):
                    ws.cell(row=current_row, column=col, value=cell_value)
                    ws.cell(row=current_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
                current_row += 1
   
    current_row += 1
   
    # Technical Evaluation Process
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = "TECHNICAL EVALUATION PROCESS"
    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{current_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
   
    process_match = re.search(r'## Technical Evaluation Process\s*\n(.*?)(?=---|$)', content, re.DOTALL)
    if process_match:
        process_text = process_match.group(1).strip()
        process_lines = [line.strip('- ').strip() for line in process_text.split('\n') if line.strip()]
       
        for line in process_lines:
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'] = line
            ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            current_row += 1
   
    # Set column widths and row heights
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
   
    # Set row heights for better readability
    for row in range(1, current_row):
        ws.row_dimensions[row].height = 20
   
    wb.save(excel_file)
    print(f"Excel file created: {excel_file}")
 
if __name__ == "__main__":
    import sys
   
    if len(sys.argv) < 2:
        print("Usage: python pure_tq_to_excel.py <markdown_file> [excel_file]")
        exit(1)
   
    md_file = sys.argv[1]
    excel_file = sys.argv[2] if len(sys.argv) > 2 else md_file.replace('.md', '.xlsx')
   
    if not os.path.exists(md_file):
        print(f"File not found: {md_file}")
        exit(1)
   
    create_tq_excel(md_file, excel_file)
 
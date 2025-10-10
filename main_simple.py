from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
import os
import uuid
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook

from src.pipeline.rfp_processor import RFPProcessor
from src.pipeline.utils import create_folder_structure, cleanup_temp_files
from src.blob_storage import blob_storage

app = FastAPI(
    title="RFP Processing Pipeline",
    description="Simple RFP processing with Azure Blob Storage",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

processor = None

def get_processor():
    global processor
    if processor is None:
        processor = RFPProcessor()
    return processor

@app.post("/process-rfp/")
async def process_rfp(file: UploadFile = File(...)):
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
   
    session_id = str(uuid.uuid4())
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
   
    try:
        session_folder = create_folder_structure(session_id, timestamp)
        pdf_path = session_folder / "input" / file.filename
        
        with open(pdf_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
       
        proc = get_processor()
        result = await proc.process_rfp(pdf_path, session_folder)
       
        # Create combined Excel
        base_path = session_folder / "excel"
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            combined_wb = Workbook()
            combined_wb.remove(combined_wb.active)
           
            excel_files = {
                'BOQ': 'boq.xlsx',
                'Prequalification': 'prequalification.xlsx', 
                'Technical_Qualification': 'technical_qualification.xlsx',
                'Summary': 'summary.xlsx',
                'Payment_Terms': 'payment_terms.xlsx'
            }
           
            for sheet_name, filename in excel_files.items():
                file_path = base_path / filename
                if file_path.exists():
                    source_wb = load_workbook(file_path)
                    source_ws = source_wb.active
                    new_ws = combined_wb.create_sheet(title=sheet_name)
                   
                    for row in source_ws.iter_rows():
                        for cell in row:
                            new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
           
            combined_wb.save(tmp_file.name)
            
            # Upload to blob storage
            download_url = blob_storage.upload_file(tmp_file.name, file.filename)
           
            cleanup_temp_files(session_folder)
            os.remove(tmp_file.name)
           
            return {"download_url": download_url, "message": "Processing completed"}
       
    except Exception as e:
        if 'session_folder' in locals():
            cleanup_temp_files(session_folder)
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")

@app.get("/files/")
async def list_files():
    """Get list of all processed files in storage"""
    try:
        files = blob_storage.list_files()
        return {"files": files, "count": len(files)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")

@app.get("/download/{filename}")
async def get_download_url(filename: str):
    """Get download URL for a specific file"""
    try:
        url = blob_storage.get_download_url(filename)
        return {"filename": filename, "download_url": url}
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"File not found: {str(e)}")

@app.get("/health/")
async def health_check():
    return {"status": "ok", "storage": "blob", "version": "1.0.0"}

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
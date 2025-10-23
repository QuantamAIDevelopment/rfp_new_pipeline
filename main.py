from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
import os
import uuid
from datetime import datetime
import asyncio
from pathlib import Path
import shutil
import tempfile
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from openpyxl.utils import get_column_letter
 
from src.pipeline.rfp_processor import RFPProcessor
from src.pipeline.utils import create_folder_structure, cleanup_temp_files
from job_store import job_store

from dotenv import load_dotenv
load_dotenv()
 
app = FastAPI(
    title="RFP Processing Pipeline",
    description="Process RFP PDFs through docling and extract structured information",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# Add CORS middleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
 
# Initialize the RFP processor (lazy loading)
processor = None
 
def get_processor():
    global processor
    if processor is None:
        processor = RFPProcessor()
    return processor
 
@app.post("/process-rfp/")
async def process_rfp(file: UploadFile = File(...)):
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
   
    job_id = str(uuid.uuid4())
    job_store.create_job(job_id, file.filename)
    
    # Save file in persistent location
    os.makedirs("/home/site/wwwroot/uploads", exist_ok=True)
    file_path = f"/home/site/wwwroot/uploads/{job_id}.pdf"
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Start background task
    asyncio.create_task(process_background(job_id, file_path, file.filename))
    
    return {"job_id": job_id, "status": "processing"}

async def process_background(job_id: str, pdf_path: str, filename: str):
    try:
        session_id = str(uuid.uuid4())
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_folder = create_folder_structure(session_id, timestamp)
        
        # Copy file to session folder
        session_pdf = session_folder / "input" / filename
        shutil.copy2(pdf_path, session_pdf)
        
        # Process
        proc = get_processor()
        await proc.process_rfp(session_pdf, session_folder)
        
        # Create result file in persistent location
        os.makedirs("/home/site/wwwroot/results", exist_ok=True)
        result_path = f"/home/site/wwwroot/results/{job_id}.xlsx"
        
        base_path = session_folder / "excel"
        combined_wb = Workbook()
        combined_wb.remove(combined_wb.active)
        
        excel_files = {
            'BOQ': 'boq.xlsx',
            'Prequalification': 'prequalification.xlsx', 
            'Technical_Qualification': 'technical_qualification.xlsx',
            'Summary': 'summary.xlsx',
            'Payment_Terms': 'payment_terms.xlsx'
        }
        
        for sheet_name, excel_filename in excel_files.items():
            file_path = base_path / excel_filename
            if file_path.exists():
                source_wb = load_workbook(file_path)
                source_ws = source_wb.active
                new_ws = combined_wb.create_sheet(title=sheet_name)
                
                # Copy all cell data with formatting
                for row in source_ws.iter_rows():
                    for cell in row:
                        new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
                            new_cell.fill = PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color, end_color=cell.fill.end_color)
                            new_cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
                            new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text)
                
                # Copy column widths
                for col in source_ws.columns:
                    column_letter = get_column_letter(col[0].column)
                    if source_ws.column_dimensions[column_letter].width:
                        new_ws.column_dimensions[column_letter].width = source_ws.column_dimensions[column_letter].width
                
                # Copy row heights
                for row_num in range(1, source_ws.max_row + 1):
                    if source_ws.row_dimensions[row_num].height:
                        new_ws.row_dimensions[row_num].height = source_ws.row_dimensions[row_num].height
        
        combined_wb.save(result_path)
        cleanup_temp_files(session_folder)
        os.remove(pdf_path)
        
        job_store.update_job(job_id, status="completed", result_file=result_path)
        
    except Exception as e:
        job_store.update_job(job_id, status="failed", error=str(e))
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
 
@app.get("/status/{job_id}")
async def get_status(job_id: str):
    job = job_store.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job

@app.get("/download/{job_id}")
async def download_result(job_id: str):
    job = job_store.get_job(job_id)
    if not job or job["status"] != "completed":
        raise HTTPException(status_code=404, detail="Result not ready")
    
    return FileResponse(
        path=job["result_file"],
        filename=f"{job['filename'].replace('.pdf', '')}_Analysis.xlsx",
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.get("/jobs/active")
async def get_active_jobs():
    """Get all currently processing jobs"""
    active_jobs = job_store.get_active_jobs()
    return {"active_jobs": active_jobs, "count": len(active_jobs)}

@app.get("/health/")
async def health_check():
    return JSONResponse(content={"status": "ok", "message": "Service is running"})

@app.get("/")
async def api_info():
    return {"message": "RFP Processing Pipeline API", "version": "1.0.0"}
 
if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port, timeout_keep_alive=1800)
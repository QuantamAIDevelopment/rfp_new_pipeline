from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
import os
import uuid
import shutil
import asyncio
import tempfile
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook

from src.pipeline.rfp_processor import RFPProcessor
from src.pipeline.utils import create_folder_structure, cleanup_temp_files
from src.blob_storage import blob_manager

app = FastAPI(
    title="RFP Processing Pipeline - Blob Storage",
    description="Enterprise RFP processing with Azure Blob Storage",
    version="2.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

processor = None

def get_processor():
    global processor
    if processor is None:
        processor = RFPProcessor()
    return processor

@app.post("/process-rfp/")
async def process_rfp(file: UploadFile = File(...)):
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files are supported")
   
    job_id = str(uuid.uuid4())
    blob_manager.create_job(job_id, file.filename)
    
    # Save file temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
        shutil.copyfileobj(file.file, tmp_file)
        temp_path = tmp_file.name
    
    # Start background processing
    asyncio.create_task(process_background(job_id, temp_path, file.filename))
    
    return {"job_id": job_id, "status": "processing"}

async def process_background(job_id: str, pdf_path: str, filename: str):
    try:
        # Process RFP
        session_id = str(uuid.uuid4())
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_folder = create_folder_structure(session_id, timestamp)
        
        session_pdf = session_folder / "input" / filename
        shutil.copy2(pdf_path, session_pdf)
        
        proc = get_processor()
        await proc.process_rfp(session_pdf, session_folder)
        
        # Create combined Excel
        base_path = session_folder / "excel"
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as result_file:
            combined_wb = Workbook()
            combined_wb.remove(combined_wb.active)
            
            excel_files = {
                'BOQ': 'boq.xlsx',
                'Prequalification': 'prequalification.xlsx', 
                'Technical_Qualification': 'technical_qualification.xlsx',
                'Summary': 'summary.xlsx',
                'Payment_Terms': 'payment_terms.xlsx'
            }
            
            for sheet_name, excel_filename in excel_files.items():
                file_path = base_path / excel_filename
                if file_path.exists():
                    source_wb = load_workbook(file_path)
                    source_ws = source_wb.active
                    new_ws = combined_wb.create_sheet(title=sheet_name)
                    
                    for row in source_ws.iter_rows():
                        for cell in row:
                            new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            
            combined_wb.save(result_file.name)
            
            # Upload to blob storage
            blob_manager.upload_result(job_id, result_file.name, filename)
            
            # Cleanup
            cleanup_temp_files(session_folder)
            os.remove(pdf_path)
            os.remove(result_file.name)
        
    except Exception as e:
        blob_manager.update_job(job_id, status="failed", error=str(e))
        if os.path.exists(pdf_path):
            os.remove(pdf_path)

@app.get("/status/{job_id}")
async def get_status(job_id: str):
    job = blob_manager.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job

@app.get("/download/{job_id}")
async def download_result(job_id: str):
    job = blob_manager.get_job(job_id)
    if not job or job["status"] != "completed":
        raise HTTPException(status_code=404, detail="Result not ready")
    
    # Redirect to blob storage URL
    return RedirectResponse(url=job["result_url"])

@app.get("/health/")
async def health_check():
    return {"status": "ok", "storage": "blob", "version": "2.0.0"}

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)